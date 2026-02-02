from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, authenticate, logout
from django.contrib import messages
from django.http import HttpResponse
from django.contrib.auth.models import User 
from django.db.models import Sum, Count, Q, Avg
from .models import SalesData, CSVUpload
from .forms import CSVUploadForm, FilterForm
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import seaborn as sns
from io import BytesIO
import base64
from datetime import datetime, timedelta
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
import json

def home(request):
    return render(request, 'analytics/home.html')

def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect('dashboard')
        messages.error(request, 'Invalid credentials')
    return render(request, 'analytics/login.html')


def signup_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')
        
        if password1 != password2:
            messages.error(request, 'Passwords do not match!')
            return redirect('signup')
        
        if User.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists!')
            return redirect('signup')
        
        try:
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password1
            )
            messages.success(request, 'Account created! Please login.')
            return redirect('login')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
            return redirect('signup')
    
    return render(request, 'analytics/signup.html')

def logout_view(request):
    logout(request)
    return redirect('login')

def calculate_insights(current_data, comparison_data=None):
    insights = []
    
    if not current_data.exists():
        return insights
    
    def to_float(value):
        """Safely convert Decimal to float"""
        if value is None:
            return 0.0
        return float(value)
    
    current_revenue = to_float(current_data.aggregate(Sum('revenue'))['revenue__sum'] or 0)
    current_leads = to_float(current_data.aggregate(Sum('leads'))['leads__sum'] or 0)
    current_conversions = to_float(current_data.aggregate(Sum('conversions'))['conversions__sum'] or 0)
    current_conv_rate = (current_conversions / current_leads * 100) if current_leads > 0 else 0

    products = current_data.values('product').annotate(
        total=Sum('revenue')
    ).order_by('-total')
    
    if products:
        best_product = products[0]
        worst_product = products[len(products)-1]
        
        best_total = to_float(best_product['total'])
        worst_total = to_float(worst_product['total'])
        
        insights.append({
            'type': 'success',
            'icon': '🏆',
            'title': 'Top Performer',
            'message': f"{best_product['product']} leads with ${best_total:,.2f} in revenue"
        })
        
        if len(products) > 1 and worst_total < best_total * 0.3:
            insights.append({
                'type': 'warning',
                'icon': '⚠️',
                'title': 'Underperformer',
                'message': f"{worst_product['product']} needs attention (${worst_total:,.2f})"
            })
    
    regions = current_data.values('region').annotate(
        total=Sum('revenue'),
        total_leads=Sum('leads'),
        total_conversions=Sum('conversions')
    ).order_by('-total')
    
    if regions:
        best_region = None
        best_conv_rate = 0
        
        for region in regions:
            leads = to_float(region['total_leads'] or 1)
            conversions = to_float(region['total_conversions'] or 0)
            conv_rate = (conversions / leads * 100) if leads > 0 else 0
            
            if conv_rate > best_conv_rate:
                best_conv_rate = conv_rate
                best_region = region
        
        if best_region:
            insights.append({
                'type': 'info',
                'icon': '🌍',
                'title': 'Best Region',
                'message': f"{best_region['region']} has highest conversion rate at {best_conv_rate:.1f}%"
            })
    
    if current_conv_rate > 30:
        insights.append({
            'type': 'success',
            'icon': '🎯',
            'title': 'Excellent Conversion',
            'message': f"Your {current_conv_rate:.1f}% conversion rate is outstanding!"
        })
    elif current_conv_rate < 20:
        insights.append({
            'type': 'warning',
            'icon': '📉',
            'title': 'Low Conversion',
            'message': f"Conversion rate of {current_conv_rate:.1f}% could be improved"
        })
    
    if comparison_data and comparison_data.exists():
        comp_revenue = to_float(comparison_data.aggregate(Sum('revenue'))['revenue__sum'] or 0)
        comp_leads = to_float(comparison_data.aggregate(Sum('leads'))['leads__sum'] or 0)
        comp_conversions = to_float(comparison_data.aggregate(Sum('conversions'))['conversions__sum'] or 0)
        
        if comp_revenue > 0:
            revenue_change = ((current_revenue - comp_revenue) / comp_revenue * 100)
            if abs(revenue_change) > 5:
                icon = '📈' if revenue_change > 0 else '📉'
                change_type = 'success' if revenue_change > 0 else 'danger'
                direction = 'up' if revenue_change > 0 else 'down'
                insights.append({
                    'type': change_type,
                    'icon': icon,
                    'title': f'Revenue {direction.title()}',
                    'message': f"Revenue {direction} {abs(revenue_change):.1f}% compared to previous period"
                })
        
        if comp_leads > 0:
            leads_change = ((current_leads - comp_leads) / comp_leads * 100)
            if abs(leads_change) > 10:
                icon = '🚀' if leads_change > 0 else '⬇️'
                change_type = 'info' if leads_change > 0 else 'warning'
                direction = 'up' if leads_change > 0 else 'down'
                insights.append({
                    'type': change_type,
                    'icon': icon,
                    'title': f'Lead Generation',
                    'message': f"Leads {direction} {abs(leads_change):.1f}% from previous period"
                })
    
    latest_date = current_data.order_by('-date').first()
    if latest_date:
        days_old = (datetime.now().date() - latest_date.date).days
        if days_old > 7:
            insights.append({
                'type': 'warning',
                'icon': '⏰',
                'title': 'Stale Data',
                'message': f"Latest data is {days_old} days old. Consider updating!"
            })
    
    return insights[:6]  # Return top 6 insights

@login_required
def dashboard(request):

    comparison_mode = request.GET.get('compare', '')
    
    sales_data = SalesData.objects.all()
    comparison_data = None
    
    preset = request.GET.get('preset', '')
    today = datetime.now().date()
    
    if preset == 'today':
        sales_data = sales_data.filter(date=today)
        if comparison_mode == 'previous':
            comparison_data = SalesData.objects.filter(date=today - timedelta(days=1))
    elif preset == 'last_7_days':
        sales_data = sales_data.filter(date__gte=today - timedelta(days=7))
        if comparison_mode == 'previous':
            comparison_data = SalesData.objects.filter(
                date__gte=today - timedelta(days=14),
                date__lt=today - timedelta(days=7)
            )
    elif preset == 'last_30_days':
        sales_data = sales_data.filter(date__gte=today - timedelta(days=30))
        if comparison_mode == 'previous':
            comparison_data = SalesData.objects.filter(
                date__gte=today - timedelta(days=60),
                date__lt=today - timedelta(days=30)
            )
    elif preset == 'this_month':
        sales_data = sales_data.filter(date__year=today.year, date__month=today.month)
        if comparison_mode == 'previous':
            last_month = today.replace(day=1) - timedelta(days=1)
            comparison_data = SalesData.objects.filter(
                date__year=last_month.year, 
                date__month=last_month.month
            )
    elif preset == 'last_month':
        last_month = today.replace(day=1) - timedelta(days=1)
        sales_data = sales_data.filter(date__year=last_month.year, date__month=last_month.month)
        if comparison_mode == 'previous':
            two_months_ago = (last_month.replace(day=1) - timedelta(days=1))
            comparison_data = SalesData.objects.filter(
                date__year=two_months_ago.year,
                date__month=two_months_ago.month
            )
    elif preset == 'this_year':
        sales_data = sales_data.filter(date__year=today.year)
        if comparison_mode == 'previous':
            comparison_data = SalesData.objects.filter(date__year=today.year - 1)
    
    filter_form = FilterForm(request.GET)
    if filter_form.is_valid():
        if filter_form.cleaned_data['start_date']:
            sales_data = sales_data.filter(date__gte=filter_form.cleaned_data['start_date'])
        if filter_form.cleaned_data['end_date']:
            sales_data = sales_data.filter(date__lte=filter_form.cleaned_data['end_date'])
        if filter_form.cleaned_data['product']:
            sales_data = sales_data.filter(product__icontains=filter_form.cleaned_data['product'])
        if filter_form.cleaned_data['region']:
            sales_data = sales_data.filter(region__icontains=filter_form.cleaned_data['region'])
    
    def to_float(value):
        if value is None:
            return 0.0
        return float(value)
    
    total_revenue = to_float(sales_data.aggregate(Sum('revenue'))['revenue__sum'] or 0)
    total_leads = to_float(sales_data.aggregate(Sum('leads'))['leads__sum'] or 0)
    total_conversions = to_float(sales_data.aggregate(Sum('conversions'))['conversions__sum'] or 0)
    conversion_rate = (total_conversions / total_leads * 100) if total_leads > 0 else 0

    comp_metrics = None
    if comparison_data and comparison_data.exists():
        comp_revenue = to_float(comparison_data.aggregate(Sum('revenue'))['revenue__sum'] or 0)
        comp_leads = to_float(comparison_data.aggregate(Sum('leads'))['leads__sum'] or 0)
        comp_conversions = to_float(comparison_data.aggregate(Sum('conversions'))['conversions__sum'] or 0)
        comp_conv_rate = (comp_conversions / comp_leads * 100) if comp_leads > 0 else 0
        
        comp_metrics = {
            'revenue': comp_revenue,
            'leads': comp_leads,
            'conversions': comp_conversions,
            'conversion_rate': comp_conv_rate,
            'revenue_change': ((total_revenue - comp_revenue) / comp_revenue * 100) if comp_revenue > 0 else 0,
            'leads_change': ((total_leads - comp_leads) / comp_leads * 100) if comp_leads > 0 else 0,
            'conversions_change': ((total_conversions - comp_conversions) / comp_conversions * 100) if comp_conversions > 0 else 0,
            'conv_rate_change': conversion_rate - comp_conv_rate,
        }
    
    top_products = sales_data.values('product').annotate(
        total=Sum('revenue')
    ).order_by('-total')[:5]
    
    top_regions = sales_data.values('region').annotate(
        total=Sum('revenue')
    ).order_by('-total')[:5]
    
    insights = calculate_insights(sales_data, comparison_data)
    
    charts = generate_plotly_charts(sales_data)
    
    context = {
        'total_revenue': total_revenue,
        'total_leads': total_leads,
        'total_conversions': total_conversions,
        'conversion_rate': round(conversion_rate, 2),
        'filter_form': filter_form,
        'charts': charts,
        'data_count': sales_data.count(),
        'top_products': top_products,
        'top_regions': top_regions,
        'current_preset': preset,
        'insights': insights,
        'comparison_mode': comparison_mode,
        'comp_metrics': comp_metrics,
    }
    return render(request, 'analytics/dashboard.html', context)

def generate_plotly_charts(queryset):
    if not queryset.exists():
        return {}
    
    df = pd.DataFrame(list(queryset.values('date', 'product', 'region', 'revenue', 'leads', 'conversions')))
    
    df['revenue'] = df['revenue'].astype(float)
    df['leads'] = df['leads'].astype(float)
    df['conversions'] = df['conversions'].astype(float)
    
    charts = {}
    
    df_grouped = df.groupby('date')['revenue'].sum().reset_index()
    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=df_grouped['date'],
        y=df_grouped['revenue'],
        mode='lines+markers',
        name='Revenue',
        line=dict(color='#3498db', width=3),
        marker=dict(size=8),
        hovertemplate='<b>Date:</b> %{x}<br><b>Revenue:</b> $%{y:,.2f}<extra></extra>'
    ))
    fig.update_layout(
        title='Revenue Over Time',
        xaxis_title='Date',
        yaxis_title='Revenue ($)',
        hovermode='x unified',
        height=400,
        template='plotly_white'
    )
    charts['revenue_trend'] = fig.to_html(full_html=False, include_plotlyjs='cdn')
    
    product_revenue = df.groupby('product')['revenue'].sum().sort_values(ascending=False)
    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=product_revenue.index,
        y=product_revenue.values,
        marker_color='#2ecc71',
        hovertemplate='<b>%{x}</b><br>Revenue: $%{y:,.2f}<extra></extra>'
    ))
    fig.update_layout(
        title='Revenue by Product',
        xaxis_title='Product',
        yaxis_title='Revenue ($)',
        height=400,
        template='plotly_white'
    )
    charts['product_revenue'] = fig.to_html(full_html=False, include_plotlyjs='cdn')
    
    fig = go.Figure()
    fig.add_trace(go.Pie(
        labels=product_revenue.index,
        values=product_revenue.values,
        hovertemplate='<b>%{label}</b><br>Revenue: $%{value:,.2f}<br>Share: %{percent}<extra></extra>',
        textposition='inside',
        textinfo='percent+label'
    ))
    fig.update_layout(
        title='Revenue Distribution by Product',
        height=400,
        template='plotly_white'
    )
    charts['product_pie'] = fig.to_html(full_html=False, include_plotlyjs='cdn')
    
    region_data = df.groupby('region').agg({'leads': 'sum', 'conversions': 'sum'})
    region_data['conv_rate'] = (region_data['conversions'] / region_data['leads'] * 100).fillna(0)
    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=region_data.index,
        y=region_data['conv_rate'],
        marker_color='#e74c3c',
        hovertemplate='<b>%{x}</b><br>Conversion Rate: %{y:.2f}%<extra></extra>'
    ))
    fig.update_layout(
        title='Conversion Rate by Region',
        xaxis_title='Region',
        yaxis_title='Conversion Rate (%)',
        height=400,
        template='plotly_white'
    )
    charts['region_conversion'] = fig.to_html(full_html=False, include_plotlyjs='cdn')
    
    region_revenue = df.groupby('region')['revenue'].sum()
    fig = go.Figure()
    fig.add_trace(go.Pie(
        labels=region_revenue.index,
        values=region_revenue.values,
        hovertemplate='<b>%{label}</b><br>Revenue: $%{value:,.2f}<br>Share: %{percent}<extra></extra>',
        textposition='inside',
        textinfo='percent+label'
    ))
    fig.update_layout(
        title='Revenue Distribution by Region',
        height=400,
        template='plotly_white'
    )
    charts['region_pie'] = fig.to_html(full_html=False, include_plotlyjs='cdn')
    
    product_data = df.groupby('product').agg({'leads': 'sum', 'conversions': 'sum'})
    fig = go.Figure()
    fig.add_trace(go.Bar(
        name='Leads',
        x=product_data.index,
        y=product_data['leads'],
        marker_color='#3498db'
    ))
    fig.add_trace(go.Bar(
        name='Conversions',
        x=product_data.index,
        y=product_data['conversions'],
        marker_color='#2ecc71'
    ))
    fig.update_layout(
        title='Leads vs Conversions by Product',
        xaxis_title='Product',
        yaxis_title='Count',
        barmode='group',
        height=400,
        template='plotly_white'
    )
    charts['leads_conversions'] = fig.to_html(full_html=False, include_plotlyjs='cdn')
    
    return charts

@login_required
def upload_csv(request):
    if request.method == 'POST':
        form = CSVUploadForm(request.POST, request.FILES)
        if form.is_valid():
            csv_file = request.FILES['file']
            
            try:
                df = pd.read_csv(csv_file)
                
                required_cols = ['date', 'product', 'region', 'revenue', 'leads', 'conversions']
                if not all(col in df.columns for col in required_cols):
                    messages.error(request, f'CSV must contain columns: {", ".join(required_cols)}')
                    return redirect('upload_csv')
                
                upload = form.save(commit=False)
                upload.user = request.user
                upload.save()
                
                rows_imported = 0
                for _, row in df.iterrows():
                    try:
                        SalesData.objects.create(
                            uploaded_by=request.user,
                            date=pd.to_datetime(row['date']).date(),
                            product=row['product'],
                            region=row['region'],
                            revenue=float(row['revenue']),
                            leads=int(row['leads']),
                            conversions=int(row['conversions'])
                        )
                        rows_imported += 1
                    except Exception as e:
                        continue
                
                upload.rows_imported = rows_imported
                upload.save()
                
                messages.success(request, f'Successfully imported {rows_imported} rows!')
                return redirect('dashboard')
                
            except Exception as e:
                messages.error(request, f'Error processing CSV: {str(e)}')
                return redirect('upload_csv')
    else:
        form = CSVUploadForm()
    
    return render(request, 'analytics/upload.html', {'form': form})

@login_required
def export_report(request):
    sales_data = SalesData.objects.all()
    
    filter_form = FilterForm(request.GET)
    if filter_form.is_valid():
        if filter_form.cleaned_data['start_date']:
            sales_data = sales_data.filter(date__gte=filter_form.cleaned_data['start_date'])
        if filter_form.cleaned_data['end_date']:
            sales_data = sales_data.filter(date__lte=filter_form.cleaned_data['end_date'])
        if filter_form.cleaned_data['product']:
            sales_data = sales_data.filter(product__icontains=filter_form.cleaned_data['product'])
        if filter_form.cleaned_data['region']:
            sales_data = sales_data.filter(region__icontains=filter_form.cleaned_data['region'])
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="sales_report_{datetime.now().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Date', 'Product', 'Region', 'Revenue', 'Leads', 'Conversions', 'Conversion Rate'])
    
    for sale in sales_data:
        conv_rate = (sale.conversions / sale.leads * 100) if sale.leads > 0 else 0
        writer.writerow([
            sale.date,
            sale.product,
            sale.region,
            float(sale.revenue),
            sale.leads,
            sale.conversions,
            f"{conv_rate:.2f}%"
        ])
    
    return response

@login_required
def export_excel(request):
    sales_data = SalesData.objects.all()
    
    filter_form = FilterForm(request.GET)
    if filter_form.is_valid():
        if filter_form.cleaned_data['start_date']:
            sales_data = sales_data.filter(date__gte=filter_form.cleaned_data['start_date'])
        if filter_form.cleaned_data['end_date']:
            sales_data = sales_data.filter(date__lte=filter_form.cleaned_data['end_date'])
        if filter_form.cleaned_data['product']:
            sales_data = sales_data.filter(product__icontains=filter_form.cleaned_data['product'])
        if filter_form.cleaned_data['region']:
            sales_data = sales_data.filter(region__icontains=filter_form.cleaned_data['region'])
    
    def to_float(value):
        if value is None:
            return 0.0
        return float(value)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    
    headers = ['Date', 'Product', 'Region', 'Revenue', 'Leads', 'Conversions', 'Conversion Rate']
    ws.append(headers)
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    for sale in sales_data:
        conv_rate = (sale.conversions / sale.leads * 100) if sale.leads > 0 else 0
        ws.append([
            sale.date.strftime('%Y-%m-%d'),
            sale.product,
            sale.region,
            to_float(sale.revenue),
            sale.leads,
            sale.conversions,
            f"{conv_rate:.2f}%"
        ])
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    ws2 = wb.create_sheet("Summary")
    ws2['A1'] = "Metric"
    ws2['B1'] = "Value"
    ws2['A1'].font = header_font
    ws2['B1'].font = header_font
    
    total_revenue = to_float(sales_data.aggregate(Sum('revenue'))['revenue__sum'] or 0)
    total_leads = to_float(sales_data.aggregate(Sum('leads'))['leads__sum'] or 0)
    total_conversions = to_float(sales_data.aggregate(Sum('conversions'))['conversions__sum'] or 0)
    conversion_rate = (total_conversions / total_leads * 100) if total_leads > 0 else 0
    
    ws2['A2'] = "Total Revenue"
    ws2['B2'] = total_revenue
    ws2['A3'] = "Total Leads"
    ws2['B3'] = total_leads
    ws2['A4'] = "Total Conversions"
    ws2['B4'] = total_conversions
    ws2['A5'] = "Conversion Rate"
    ws2['B5'] = f"{conversion_rate:.2f}%"
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="sales_report_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    
    return response

@login_required
def delete_data(request):
    if request.method == 'POST':
        delete_type = request.POST.get('delete_type')
        
        if delete_type == 'all':
            count = SalesData.objects.all().count()
            SalesData.objects.all().delete()
            messages.success(request, f'Deleted all {count} records!')
        elif delete_type == 'filtered':
            sales_data = SalesData.objects.all()
            filter_form = FilterForm(request.POST)
            if filter_form.is_valid():
                if filter_form.cleaned_data['start_date']:
                    sales_data = sales_data.filter(date__gte=filter_form.cleaned_data['start_date'])
                if filter_form.cleaned_data['end_date']:
                    sales_data = sales_data.filter(date__lte=filter_form.cleaned_data['end_date'])
                if filter_form.cleaned_data['product']:
                    sales_data = sales_data.filter(product__icontains=filter_form.cleaned_data['product'])
                if filter_form.cleaned_data['region']:
                    sales_data = sales_data.filter(region__icontains=filter_form.cleaned_data['region'])
            
            count = sales_data.count()
            sales_data.delete()
            messages.success(request, f'Deleted {count} filtered records!')
        
        return redirect('dashboard')
    
    return render(request, 'analytics/delete_confirm.html')
